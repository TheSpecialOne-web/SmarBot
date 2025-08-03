from copy import deepcopy
import html
from typing import Any

from azure.ai.documentintelligence.models import DocumentTable as DocumentIntelligenceTable
from azure.ai.formrecognizer import DocumentTable as FormRecognizerTable
from bs4 import BeautifulSoup
from docx.table import Table as DocxTable
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from pydantic import BaseModel, Field, NonNegativeInt, StrictBool, StrictStr, field_validator, model_validator
import tiktoken
from xlrd.sheet import Sheet


class DocumentUploadQueue(BaseModel):
    tenant_id: int
    bot_id: int
    document_id: int


class CreateEmbeddingsQueue(BaseModel):
    tenant_id: int
    bot_id: int
    document_id: int


class CalculateStorageUsageQueue(BaseModel):
    tenant_id: int
    bot_id: int
    document_id: int


class SyncDocumentPathQueue(BaseModel):
    tenant_id: int
    bot_id: int
    document_folder_id: str
    document_ids: list[int]


DEFAULT_TABLE_CHUNK_SIZE = 1500
DEFAULT_ENCODING_NAME = "o200k_base"
MAX_HEADER_WIDTH = 3

# 計算量、メモリ使用量を抑えるため、テーブルの最大サイズを制限
MAX_ROW_NUMBER_FOR_SHEET = 1000
MAX_COL_NUMBER_FOR_SHEET = 1000

Coord = tuple[int, int]


class ParentCell(BaseModel):
    text: StrictStr = ""
    row_span: NonNegativeInt = 1
    col_span: NonNegativeInt = 1
    is_header: StrictBool = False

    @field_validator("text", mode="after")
    @classmethod
    def clean_up_text(cls, value: str) -> str:
        return value.replace(":selected:", "").replace(":unselected:", "").strip()

    @property
    def is_merged(self) -> bool:
        return self.row_span >= 2 or self.col_span >= 2

    @property
    def is_blank(self) -> bool:
        return self.text.strip() == ""


class ChildCell(BaseModel):
    parent_coord: Coord


Cell = ParentCell | ChildCell


class Table(BaseModel):
    matrix: list[list[Cell]]
    caption: StrictStr | None = None
    header_row_idxs: list[int] = Field(default_factory=list)

    @model_validator(mode="after")  # type: ignore
    def validate_list(self) -> "Table":
        if not all(len(row) == len(self.matrix[0]) for row in self.matrix):
            raise ValueError("All rows should have the same length")
        return self

    @model_validator(mode="after")  # type: ignore
    def validate_child(self) -> "Table":
        for row in self.matrix:
            for cell in row:
                if isinstance(cell, ParentCell):
                    continue
                parent_cell = self.matrix[cell.parent_coord[0]][cell.parent_coord[1]]
                if not isinstance(parent_cell, ParentCell):
                    raise ValueError("Parent cell is not found")
                if not parent_cell.is_merged:
                    raise ValueError("Parent cell is not merged cell")
        return self

    @property
    def h(self) -> int:
        return len(self.matrix)

    @property
    def w(self) -> int:
        return len(self.matrix[0])

    def _is_blank_row(self, row_idx: int) -> bool:
        if not 0 <= row_idx < self.h:
            raise ValueError(f"Invalid row index: {row_idx} input should be within (0, {self.h-1})")

        row_cells = self.matrix[row_idx]
        return all(isinstance(cell, ChildCell) or cell.is_blank for cell in row_cells)

    def _is_blank_col(self, col_idx: int) -> bool:
        if not 0 <= col_idx < self.w:
            raise ValueError(f"Invalid col index: {col_idx} input should be within (0, {self.w-1})")

        col_cells = [row[col_idx] for row in self.matrix]
        return all(isinstance(cell, ChildCell) or cell.is_blank for cell in col_cells)

    def drop_row(self, row_idx: int) -> None:
        if not 0 <= row_idx < self.h:
            raise ValueError(f"Invalid row index {row_idx} input should be within (0, {self.h-1})")

        # ParentCellの情報の更新
        revised_parent_coords = set()  # 余分に引かれるのを防ぐため
        for col_idx, cell in enumerate(self.matrix[row_idx]):
            if isinstance(cell, ParentCell):
                # 単体のセル
                if not cell.is_merged:
                    continue
                # 縦並びの結合セル
                if cell.row_span == 1:
                    continue
                # 親セルをずらす
                if row_idx + 1 < self.h:
                    self.matrix[row_idx + 1][col_idx] = ParentCell(
                        text=cell.text,
                        row_span=cell.row_span - 1,
                        col_span=cell.col_span,
                        is_header=cell.is_header,
                    )

            if isinstance(cell, ChildCell) and cell.parent_coord not in revised_parent_coords:
                parent_row_idx, parent_col_idx = cell.parent_coord
                parent_cell = self.matrix[parent_row_idx][parent_col_idx]
                if isinstance(parent_cell, ChildCell):
                    raise ValueError("Parent cell is not found")
                parent_cell.row_span -= 1
                revised_parent_coords.add(cell.parent_coord)

        # ChildCellの情報の更新（parent_coordをずらす）
        for _row_idx in range(row_idx + 1, self.h):
            for _col_idx in range(self.w):
                cell = self.matrix[_row_idx][_col_idx]
                if isinstance(cell, ChildCell) and cell.parent_coord[0] > row_idx:
                    cell.parent_coord = (cell.parent_coord[0] - 1, cell.parent_coord[1])

        self.matrix.pop(row_idx)

    def drop_col(self, col_idx: int) -> None:
        if not 0 <= col_idx < self.w:
            raise ValueError(f"Invalid col index: {col_idx} input should be within (0, {self.w-1})")

        # ParentCellの情報の更新
        revised_parent_coords = set()  # 余分に引かれるのを防ぐため
        for row_idx in range(self.h):
            cell = self.matrix[row_idx][col_idx]

            if isinstance(cell, ParentCell):
                # 単体のセル
                if not cell.is_merged:
                    continue
                # 縦並びの結合セル
                if cell.col_span == 1:
                    continue
                # 親セルをずらす
                if col_idx + 1 < self.w:
                    self.matrix[row_idx][col_idx + 1] = ParentCell(
                        text=cell.text,
                        row_span=cell.row_span,
                        col_span=cell.col_span - 1,
                        is_header=cell.is_header,
                    )

            if isinstance(cell, ChildCell) and cell.parent_coord not in revised_parent_coords:
                parent_row_idx, parent_col_idx = cell.parent_coord
                parent_cell = self.matrix[parent_row_idx][parent_col_idx]
                if isinstance(parent_cell, ChildCell):
                    raise ValueError("Parent cell is not found")
                parent_cell.col_span -= 1
                revised_parent_coords.add(cell.parent_coord)

        # ChildCellの情報の更新（parent_coordをずらす）
        for _row_idx in range(self.h):
            for _col_idx in range(col_idx + 1, self.w):
                cell = self.matrix[_row_idx][_col_idx]
                if isinstance(cell, ChildCell) and cell.parent_coord[1] > col_idx:
                    cell.parent_coord = (cell.parent_coord[0], cell.parent_coord[1] - 1)

        for row in self.matrix:
            row.pop(col_idx)

    def drop_blank_rows_and_cols(self) -> None:
        for row_idx in reversed(range(self.h)):
            if self._is_blank_row(row_idx):
                self.drop_row(row_idx)
        if len(self.matrix) == 0:
            return
        for col_idx in reversed(range(self.w)):
            if self._is_blank_col(col_idx):
                self.drop_col(col_idx)

    def drop_non_header_rows(self, row_idx_start: int, row_idx_end: int, max_header_row_count: int = 3) -> None:
        undroppable_header_row_idxs = [
            _row_idx
            for _row_idx in range(
                row_idx_start,
                row_idx_end,
            )
            if _row_idx in self.header_row_idxs
        ]
        # ヘッダー行が多すぎて過剰に追加されてしまう場合のため、追加するヘッダーの数をmax_header_row_countに制限
        if len(undroppable_header_row_idxs) > max_header_row_count:
            undroppable_header_row_idxs = undroppable_header_row_idxs[: max_header_row_count - 1] + [
                undroppable_header_row_idxs[-1]
            ]
        for row_idx in reversed(range(row_idx_start, row_idx_end)):
            if row_idx in undroppable_header_row_idxs:
                continue
            self.drop_row(row_idx)

    def _contains_merged_cell(self) -> bool:
        return any(isinstance(cell, ChildCell) for row in self.matrix for cell in row)

    def to_text(self, row_idx_start: int = 0, row_idx_end: int | None = None) -> str:
        if row_idx_end is None:
            row_idx_end = len(self.matrix)
        matrix = deepcopy(self.matrix[0:row_idx_end])
        table = Table(matrix=matrix, caption=self.caption, header_row_idxs=self.header_row_idxs)
        table.drop_non_header_rows(0, row_idx_start)

        table.drop_blank_rows_and_cols()
        if table._contains_merged_cell():
            return table._to_html()
        return table._to_md()

    def _to_html(self) -> str:
        try:
            soup = BeautifulSoup("", "html.parser")
            table_soup = soup.new_tag("table")

            if self.caption:
                caption_tag = soup.new_tag("caption")
                caption_tag.append(self.caption)
                table_soup.append(caption_tag)

            for row in self.matrix:
                tr_tag = soup.new_tag("tr")
                table_soup.append(tr_tag)
                for cell in row:
                    # 子セルはスキップ
                    if isinstance(cell, ChildCell):
                        continue

                    # 親セルもしくは単一セルの場合
                    attrs: dict[str, str] = {}
                    if cell.col_span > 1:
                        attrs["colSpan"] = str(cell.col_span)
                    if cell.row_span > 1:
                        attrs["rowSpan"] = str(cell.row_span)

                    if cell.is_header:
                        new_tag = soup.new_tag("th", attrs=attrs)
                    else:
                        new_tag = soup.new_tag("td", attrs=attrs)
                    new_tag.append(html.escape(cell.text))
                    tr_tag.append(new_tag)

                table_soup.append(tr_tag)
            return table_soup.prettify() + "\n"
        finally:
            soup.decompose()

    def _to_md(self) -> str:
        def get_cell_text(cell: Cell) -> str:
            if isinstance(cell, ParentCell):
                return cell.text
            raise ValueError("Child cell is not allowed")

        matrix_data = [[get_cell_text(cell) for cell in row] for row in self.matrix]
        df = pd.DataFrame(matrix_data[1:], columns=matrix_data[0])

        # データフレームの各セルに対して改行をエスケープ
        def _escape_newlines(text):
            return text.replace("\n", "<br>").replace("|", "\\|")

        df = df.apply(lambda x: x.apply(_escape_newlines))

        if self.caption:
            md_text = f"{df.to_markdown(index=False)}\nTable: {self.caption}"
        else:
            md_text = df.to_markdown(index=False)
        return f"<table>\n{md_text}\n</table>\n"

    def _get_token_count(self, string: str) -> int:
        encoding = tiktoken.get_encoding(DEFAULT_ENCODING_NAME)
        num_tokens = len(encoding.encode(string))
        return num_tokens

    def to_chunks(self, table_chunk_size: int | None = None) -> list[str]:
        table_chunk_size = table_chunk_size or DEFAULT_TABLE_CHUNK_SIZE

        chunks: list[str] = []

        row_idx_start = 0
        row_idx_end = 1
        while row_idx_end <= self.h:
            text = self.to_text(row_idx_start, row_idx_end)
            if row_idx_end == self.h:
                chunks.append(text)
                break

            if self._get_token_count(text) <= table_chunk_size:
                row_idx_end += 1
                continue

            # 一列でチャンクサイズを超える場合、一行で追加
            if row_idx_end - row_idx_start > 1:
                row_idx_end = row_idx_end - 1

            # cut_idxの行に子セルがあり、その親セルがチャンク内にある場合、親セルを新たに設定
            for _col_idx in range(self.w):
                cell = self.matrix[row_idx_end][_col_idx]
                if isinstance(cell, ParentCell):
                    continue
                if cell.parent_coord[0] == row_idx_end:
                    continue

                parent_row_idx, _col_idx = cell.parent_coord
                parent_cell = self.matrix[parent_row_idx][_col_idx]

                if not isinstance(parent_cell, ParentCell):
                    raise ValueError("Parent cell is not found")

                parent_cell_row_span = row_idx_end - parent_row_idx
                new_parent_cell_row_span = parent_cell.row_span - parent_cell_row_span

                # 区切れた先頭のセルを親セルに設定
                if isinstance(parent_cell, ChildCell):
                    raise ValueError("Parent cell is not found")
                new_parent_cell = ParentCell(
                    text=parent_cell.text,
                    row_span=new_parent_cell_row_span,
                    col_span=parent_cell.col_span,
                    is_header=parent_cell.is_header,
                )
                self.matrix[row_idx_end][_col_idx] = new_parent_cell

                # チャンク内の親セルのrow_spanを更新
                parent_cell.row_span = parent_cell_row_span

                # 子セルのparent_coordを設定
                for child_row_idx in range(row_idx_end, row_idx_end + new_parent_cell.row_span):
                    for child_col_idx in range(_col_idx, _col_idx + new_parent_cell.col_span):
                        if (child_row_idx, child_col_idx) == (row_idx_end, _col_idx):
                            continue
                        cell = self.matrix[child_row_idx][child_col_idx]
                        if isinstance(cell, ParentCell):
                            raise ValueError("Child cell is not found")
                        cell.parent_coord = (row_idx_end, _col_idx)

            chunks.append(self.to_text(row_idx_start, row_idx_end))
            row_idx_start = row_idx_end
            row_idx_end += 1

        return chunks

    @classmethod
    def from_document_intelligence_table(cls, table_di: FormRecognizerTable | DocumentIntelligenceTable) -> "Table":
        table = cls(
            matrix=[[ParentCell() for _ in range(table_di.column_count)] for _ in range(table_di.row_count)],
        )

        for cell_di in table_di.cells:
            row_span = cell_di.row_span or 1
            col_span = cell_di.column_span or 1

            table.matrix[cell_di.row_index][cell_di.column_index] = ParentCell(
                text=cell_di.content,
                row_span=row_span,
                col_span=col_span,
                is_header=cell_di.kind == "columnHeader" or cell_di.kind == "rowHeader",
            )

            for child_row_idx in range(cell_di.row_index, cell_di.row_index + row_span):
                for child_col_idx in range(cell_di.column_index, cell_di.column_index + col_span):
                    if (child_row_idx, child_col_idx) == (
                        cell_di.row_index,
                        cell_di.column_index,
                    ):
                        continue

                    table.matrix[child_row_idx][child_col_idx] = ChildCell(
                        parent_coord=(cell_di.row_index, cell_di.column_index)
                    )

        # ヘッダーブロックのインデックスを取得
        table.header_row_idxs = [
            row_idx
            for row_idx, row in enumerate(table.matrix)
            if any(isinstance(cell, ParentCell) and cell.is_header for cell in row)
        ]
        return table

    @classmethod
    def from_docx_table(cls, table_docx: DocxTable) -> "Table":
        table = cls(
            matrix=[[ParentCell() for _ in range(len(table_docx.columns))] for _ in range(len(table_docx.rows))],
        )

        for row_idx, row in enumerate(table_docx.rows):
            for col_idx, cell in enumerate(row.cells):
                if isinstance(table.matrix[row_idx][col_idx], ChildCell):
                    continue

                col_span = cell._element.right - cell._element.left
                row_span = cell._element.bottom - cell._element.top

                table.matrix[row_idx][col_idx] = ParentCell(
                    text=cell.text or "",
                    col_span=col_span,
                    row_span=row_span,
                )
                if col_span == 1 and row_span == 1:
                    continue

                for child_row_idx in range(row_idx, row_idx + row_span):
                    for child_col_idx in range(col_idx, col_idx + col_span):
                        if (child_row_idx, child_col_idx) == (row_idx, col_idx):
                            continue
                        table.matrix[child_row_idx][child_col_idx] = ChildCell(parent_coord=(row_idx, col_idx))
        table.header_row_idxs = [0]
        for cell in table.matrix[0]:
            if isinstance(cell, ParentCell) and cell.row_span > 1:
                for _row_idx in range(row_idx, row_idx + min(cell.row_span, MAX_HEADER_WIDTH)):
                    table.header_row_idxs.append(_row_idx)
        return table

    @classmethod
    def from_xlsx_table(cls, table_xlsx: Worksheet, sheet_name: str) -> "Table":
        # テーブルの値の存在する範囲を取得
        # 注意： openpyxlでのテーブルは1から数え始めるのでopenpyxlの関数のI/Oでは±1をして対応する
        row_idx_bottom, col_idx_right, row_idx_top, col_idx_left = (
            table_xlsx.max_row - 1,
            table_xlsx.max_column - 1,
            table_xlsx.min_row - 1,
            table_xlsx.min_column - 1,
        )

        def is_nulllike(value: Any) -> bool:
            return value is None or str(value).strip() == ""

        existing_row_idx_bottom = next(
            (
                _row
                for _row in range(row_idx_bottom, row_idx_top - 1, -1)
                if any(
                    not is_nulllike(table_xlsx.cell(row=_row + 1, column=_col + 1).value)
                    for _col in range(col_idx_left, col_idx_right + 1)
                )
            ),
            -1,
        )
        existing_col_idx_right = next(
            (
                _col
                for _col in range(col_idx_right, col_idx_left - 1, -1)
                if any(
                    not is_nulllike(table_xlsx.cell(row=_row + 1, column=_col + 1).value)
                    for _row in range(row_idx_top, existing_row_idx_bottom + 1)
                )
            ),
            -1,
        )
        existing_row_idx_top = next(
            (
                _row
                for _row in range(row_idx_top, row_idx_bottom + 1)
                if any(
                    not is_nulllike(table_xlsx.cell(row=_row + 1, column=_col + 1).value)
                    for _col in range(col_idx_left, existing_col_idx_right + 1)
                )
            ),
            -1,
        )
        existing_col_idx_left = next(
            (
                _col
                for _col in range(col_idx_left, existing_col_idx_right + 1)
                if any(
                    not is_nulllike(table_xlsx.cell(row=_row + 1, column=_col + 1).value)
                    for _row in range(existing_row_idx_top, existing_row_idx_bottom + 1)
                )
            ),
            -1,
        )
        if (
            existing_col_idx_left == -1
            or existing_col_idx_right == -1
            or existing_row_idx_top == -1
            or existing_row_idx_bottom == -1
        ):
            raise ValueError("Table is empty")

        if (
            existing_col_idx_right - existing_col_idx_left + 1 > MAX_COL_NUMBER_FOR_SHEET
            or existing_row_idx_bottom - existing_row_idx_top + 1 > MAX_ROW_NUMBER_FOR_SHEET
        ):
            raise ValueError("Too large table")
        table = cls(
            matrix=[
                [ParentCell(text=str(cell.value) if cell.value is not None else "") for cell in row]
                for row in table_xlsx.iter_rows(
                    min_row=existing_row_idx_top + 1,
                    min_col=existing_col_idx_left + 1,
                    max_row=existing_row_idx_bottom + 1,
                    max_col=existing_col_idx_right + 1,
                )
            ],
            caption=sheet_name,
        )

        # 結合セルの設定
        for merged_range in table_xlsx.merged_cells.ranges:
            col_idx_left, row_idx_top, col_idx_right, row_idx_bottom = [
                idx - 1 for idx in range_boundaries(str(merged_range))
            ]
            if (
                col_idx_left > existing_col_idx_right
                or row_idx_top > existing_row_idx_bottom
                or col_idx_right < existing_col_idx_left
                or row_idx_bottom < existing_row_idx_top
            ):
                continue
            # 領域内に結合セルが収まるように範囲を調整、領域内でのインデックスに数え直す
            merged_cell_left_idx = max(col_idx_left, existing_col_idx_left) - existing_col_idx_left
            merged_cell_right_idx = min(col_idx_right, existing_col_idx_right) - existing_col_idx_left
            merged_cell_top_idx = max(row_idx_top, existing_row_idx_top) - existing_row_idx_top
            merged_cell_bottom_idx = min(row_idx_bottom, existing_row_idx_bottom) - existing_row_idx_top

            if isinstance(table.matrix[merged_cell_top_idx][merged_cell_left_idx], ChildCell):
                raise ValueError("Parent cell is not found")

            table.matrix[merged_cell_top_idx][merged_cell_left_idx] = ParentCell(
                text=table.matrix[merged_cell_top_idx][merged_cell_left_idx].text,
                row_span=merged_cell_bottom_idx - merged_cell_top_idx + 1,
                col_span=merged_cell_right_idx - merged_cell_left_idx + 1,
            )

            for child_row_idx in range(merged_cell_top_idx, merged_cell_bottom_idx + 1):
                for child_col_idx in range(merged_cell_left_idx, merged_cell_right_idx + 1):
                    if (child_row_idx, child_col_idx) == (merged_cell_top_idx, merged_cell_left_idx):
                        continue
                    table.matrix[child_row_idx][child_col_idx] = ChildCell(
                        parent_coord=(merged_cell_top_idx, merged_cell_left_idx)
                    )

        table.header_row_idxs = []
        for row_idx, row in enumerate(table.matrix):
            if all(
                (isinstance(cell, ParentCell) and cell.text != "")
                or (isinstance(cell, ChildCell) and cell.parent_coord[0] == row_idx)
                for cell in row
            ):
                # タイトル除け
                if len([cell for cell in row if isinstance(cell, ParentCell) and cell.text != ""]) == 1:
                    continue
                table.header_row_idxs.append(row_idx)
                for cell in row:
                    if isinstance(cell, ParentCell) and cell.row_span > 1:
                        for _row_idx in range(row_idx, row_idx + min(cell.row_span, MAX_HEADER_WIDTH)):
                            table.header_row_idxs.append(_row_idx)
                break

        for row_idx in table.header_row_idxs:
            for cell in table.matrix[row_idx]:
                if isinstance(cell, ParentCell):
                    cell.is_header = True
        return table

    @classmethod
    def from_xls_table(cls, sheet: Sheet, sheet_name: str) -> "Table":
        existing_max_row, existing_max_col = 0, 0
        for _row in range(sheet.nrows - 1, -1, -1):
            if any(sheet.cell_value(_row, _col) not in (None, "") for _col in range(sheet.ncols)):
                existing_max_row = _row + 1
                break
        for _col in range(sheet.ncols - 1, -1, -1):
            if any(sheet.cell_value(_row, _col) not in (None, "") for _row in range(sheet.nrows)):
                existing_max_col = _col + 1
                break
        if existing_max_row == 0 or existing_max_col == 0:
            raise ValueError(f"Sheet '{sheet_name}' is empty.")
        if existing_max_row > 1000 or existing_max_col > 1000:
            raise ValueError("Too large table")
        table = cls(
            matrix=[
                [
                    ParentCell(
                        text=(
                            str(sheet.cell_value(row_idx, col_idx))
                            if sheet.cell_value(row_idx, col_idx) is not None
                            else ""
                        )
                    )
                    for col_idx in range(sheet.ncols)
                ]
                for row_idx in range(sheet.nrows)
            ],
            caption=sheet_name,
        )

        for merged_range in sheet.merged_cells:
            min_row, max_row, min_col, max_col = merged_range
            if isinstance(table.matrix[min_row][min_col], ChildCell):
                raise ValueError("Parent cell is not found")

            table.matrix[min_row][min_col] = ParentCell(
                text=table.matrix[min_row][min_col].text,
                row_span=max_row - min_row,
                col_span=max_col - min_col,
            )

            for child_row_idx in range(min_row, max_row):
                for child_col_idx in range(min_col, max_col):
                    if (child_row_idx, child_col_idx) == (min_row, min_col):
                        continue
                    table.matrix[child_row_idx][child_col_idx] = ChildCell(parent_coord=(min_row, min_col))

        return table
