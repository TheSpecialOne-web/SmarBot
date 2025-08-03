# NMSデータ、安全サイトデータ、TIOSデータのExcelファイルを手動で取り込む前処理を行うモジュール
import os

import openpyxl as px
from openpyxl.styles import Font, PatternFill
import pandas as pd


def split_NMS_files(original_path: str, base_directory: str) -> None:
    """
    1ファイルに大量の件名があるNMSデータを、管理番号ごとに分割してExcelファイルとして保存する関数
    Args:
        original_path (str): 元のExcelファイルのパス
        base_directory (str): 分割後のExcelファイルを保存するディレクトリのパス
    """
    # 分割するための列名
    ID_COL_NAME = "管理番号"
    CASE_COL_NAME = "障害件名"
    BRANCH_COL_NAME = "作成者機関名(支店)"
    DATE_COL_NAME = "受付日時"
    BRANCH_REPLACE_WORDS = ["支社", "支店"]

    # フォーマットを整えるための設定
    DEFAULT_COL_WIDTH = 10
    MID_COL_WIDTH = 25
    LONG_COL_WIDTH = 75
    MID_COL_NAMES = ["管理番号", "対応区分", "対応レベル", "障害発生状況", "障害区間検討結果", "対応日時"]
    LONG_COL_NAMES = ["障害件名", "対応履歴内容"]
    RED_COL_NAMES = [
        "管理番号",
        "障害件名",
        "作成者機関名(所属)",
        "対応区分",
        "対応レベル",
        "障害発生日時",
        "障害/申告内容",
        "障害区間検討結果",
        "復旧指示内容",
        "障害処置結果",
        "対応履歴内容",
    ]

    df = pd.read_excel(original_path)
    df_list = [df[1] for df in df.groupby(ID_COL_NAME)]
    for df in df_list:
        file_name = f"{df[ID_COL_NAME].iloc[0]}_{df[CASE_COL_NAME].iloc[0]}.xlsx"
        file_name = file_name.replace("/", "_")
        branch = str(df[BRANCH_COL_NAME].iloc[0])
        for word in BRANCH_REPLACE_WORDS:
            branch = branch.replace(word, "")
        year = str(df[DATE_COL_NAME].iloc[0])[:4] + "年度"
        new_dir_name = f"{base_directory}/{branch}/{year}"
        full_path = f"{new_dir_name}/{file_name}"
        if os.path.exists(full_path):
            print(f"ファイル{full_path}はすでに存在しています")
            continue
        os.makedirs(new_dir_name, exist_ok=True)
        df.to_excel(full_path, index=False)

        # pandasでsaveするだけではフォーマットが崩れてしまうので、openpyxlで再度フォーマットを整える
        wb = px.load_workbook(full_path)
        ws = wb.active
        for col in ws.columns:
            col_alphabet = col[0].column_letter
            col_name = col[0].value
            col[0].alignment = px.styles.Alignment(horizontal="left")
            if col_name in MID_COL_NAMES:
                ws.column_dimensions[col_alphabet].width = MID_COL_WIDTH
            elif col_name in LONG_COL_NAMES:
                ws.column_dimensions[col_alphabet].width = LONG_COL_WIDTH
            else:
                ws.column_dimensions[col_alphabet].width = DEFAULT_COL_WIDTH
            if col_name in RED_COL_NAMES:
                for cell in col:
                    cell.font = px.styles.fonts.Font(color="FF0000")
        wb.save(full_path)
        print(f"ファイルを保存しました:{full_path}")

    print(f"分割が完了しました:{len(df_list)}ファイル")


#  安全関係のファイルを分割する関数を実装
def split_safety_files(original_path: str, base_directory: str) -> None:
    """
    1ファイルに大量の件名がある安全サイトの資料検索のデータを、分割してExcelファイルとして保存する関数
    Args:
        original_path (str): 元のExcelファイルのパス
        base_directory (str): 分割後のExcelファイルを保存するディレクトリのパス
    """
    # 分割するための列名
    CASE_COL_NAME = "件名"
    BRANCH_COL_NAME = "管轄支社"
    DATE_COL_NAME = "日付"

    BRANCH_DELETE_WORDS = ["（送配電）", "（九電）"]

    # フォーマットを整えるための設定
    DEFAULT_COL_WIDTH = 10
    MID_COL_WIDTH = 15
    LONG_COL_WIDTH = 50
    MID_COL_NAMES = ["日付", "場所（具体的な場所）", "管轄支社", "作業種別", "場所種別"]
    LONG_COL_NAMES = ["件名", "概要"]
    RED_COL_NAMES = ["管轄支社", "作業種別", "場所種別"]

    df = pd.read_excel(original_path)
    df_without_col_1 = df.iloc[:, 1:]

    case_df_list = [df[1] for df in df_without_col_1.groupby(CASE_COL_NAME)]

    for case_df in case_df_list:
        file_name = f"{case_df[CASE_COL_NAME].iloc[0]}.xlsx"
        file_name = file_name.replace("/", "_")
        branch = str(case_df[BRANCH_COL_NAME].iloc[0])
        for del_word in BRANCH_DELETE_WORDS:
            branch = branch.replace(del_word, "")
        if branch == "/":
            branch = "その他"
        year = str(case_df[DATE_COL_NAME].iloc[0])[:4] + "年度"
        new_dir_name = f"{base_directory}/{branch}/{year}"
        full_path = f"{new_dir_name}/{file_name}"
        if os.path.exists(full_path):
            print(f"ファイル{full_path}はすでに存在しています")
            continue
        os.makedirs(new_dir_name, exist_ok=True)
        case_df.to_excel(full_path, index=False)

        # pandasでsaveするだけではフォーマットが崩れてしまうので、openpyxlで再度フォーマットを整える
        wb = px.load_workbook(full_path)
        ws = wb.active
        for col in ws.columns:
            col_alphabet = col[0].column_letter
            col_name = col[0].value
            col[0].alignment = px.styles.Alignment(horizontal="left")
            col[0].font = Font(bold=True)
            gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            col[0].fill = gray_fill
            if col_name in MID_COL_NAMES:
                ws.column_dimensions[col_alphabet].width = MID_COL_WIDTH
            elif col_name in LONG_COL_NAMES:
                ws.column_dimensions[col_alphabet].width = LONG_COL_WIDTH
            else:
                ws.column_dimensions[col_alphabet].width = DEFAULT_COL_WIDTH
            if col_name in RED_COL_NAMES:
                col[1].font = px.styles.fonts.Font(color="FF0000")

        wb.save(full_path)
        print(f"ファイルを保存しました:{full_path}")

    print(f"分割が完了しました:{len(case_df_list)}ファイル")


def split_TIOS_files(original_path: str, base_directory: str) -> None:  # TODO: TIOSのファイルを分割する関数を実装
    return
